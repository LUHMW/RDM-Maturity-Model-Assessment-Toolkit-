import sys
#para leer argumentos que Excel pasa
from pathlib import Path
# Path maneja rutas de archivos de forma segura
import pandas as pd 
#pandas trabaja con tablas
from openpyxl import load_workbook
#para abrir archivos .xslm
from owlready2 import get_ontology, World
from owlready2 import Imp
import tempfile 
import xlwings as xw 
from owlready2 import sync_reasoner_pellet
from owlready2 import onto_path
import os


#Das Blatt „ImportedData” aus Excel wird gelesen, ohne die Kopfzeile zu berücksichtigen.
def read_imported_data(xslm_path:Path)->pd.DataFrame:
    wb=load_workbook(xslm_path,data_only=True,keep_vba=True)

    #Es wird überprüft, ob "ImportedData" in Excel vorhanden ist
    if "ImportedData" not in wb.sheetnames:
        raise ValueError("ImportedData ist nicht vohanden in der Excel")
    ws=wb["ImportedData"]

    #Alle Zeilen des Blattes ohne Kopfzeile werden gelesen
    rows=list(ws.iter_rows(values_only=True))
    if not rows or len(rows)<2:
        raise ValueError("ImportedData ist leer.")
    
    header=[str(h).strip()if h is not None else "" for h in rows[0]]
    data = rows[1:]

    #DataFrame mit den Daten erstellt
    df=pd.DataFrame(data, columns=header)
    df=df.dropna(axis=1, how="all")
    return df

def write_sheets(xlsm_fullname: str, df: pd.DataFrame, first_row: int=30):

    #Datei öffnen
    xwb=xw.Book(xlsm_fullname)
    app=xwb.app

    #Wörterbuch mit den Namen der Blätter, die in der Excel-Datei gespeichert sind
    sheet_name_mapping={
        sheet.name.strip().lower():sheet 
        for sheet in xwb.sheets
    }
    
    #aktuellen Status speichern
    current_screen=app.screen_updating
    app.screen_updating=False

    try:
  
     for phase, gruppe in df.groupby("Phase", sort=False):
         
         phase_name=str(phase).strip().lower()
         #print("Escribiendo phase: ", phase_name, "filas:", len(gruppe))

         #Suche nach dem Blatt, dessen Name mit der Phase übereinstimmt
         sheet=sheet_name_mapping.get(phase_name)

         if sheet is None:
            continue

         #Letzte vom Blatt verwendete Zeile
         last_row= sheet.used_range.last_cell.row

         if last_row<first_row:
            last_row=first_row
        
         #Der Inhalt ab Zeile 30 wird gelöscht
         sheet.range((first_row,1), (last_row,4)).clear_contents()

         #Überschriften
         header=["Zielindex", "Ziel", "Praktikindex", "Praktik"]
         header_range=sheet.range((first_row,1),(first_row,4))
         header_range.value=header
         header_range.api.Font.Bold=True
         header_range.row_height=18

         

         #Nur die zum Schreiben erforderlichen Spalten extrahieren
         columns_names=gruppe[["Zielindex", "Ziel", "Praktikindex", "Praktik"]].values.tolist()

         sheet.range((first_row+1,1)).value=columns_names

         #Breite der Spalten
         sheet.range("A:A").column_width=10 #Zielindex Spalte 
         sheet.range("B:B").column_width=45 #Zieltext Spalte
         sheet.range("C:C").column_width=12 #Praktikindex Spalte
         sheet.range("D:D").column_width=55 #Praktik

         #Text anpassen, um Abbrüche zu vermeiden
         sheet.range((first_row,1),
                     (first_row+ max(len(columns_names),1),4)).api.WrapText=True


     #app.calculate()
     #print("FIN escritura")

    finally:
     app.screen_updating=current_screen





#Das Blatt "Ontology" aus Excel wird gelesen und wieder verbunden.
def read_ontology_text(xlsm: Path)-> str:
    wb=load_workbook(xlsm, data_only=True,keep_vba=True)
    if "Ontology" not in wb.sheetnames:
        raise ValueError("Die Ontology ist in der Excel nicht vorhanden.")
    ws=wb["Ontology"]

    parts= [] #Liste mit den Teilen des Owls
    for row in ws.iter_rows(min_row=2, min_col=1,max_col=1, values_only=True):
        val= row[0]
        if val:
            parts.append(str(val))
    owl_text="".join(parts).strip() #Verbindung aller Teile des Owls
    owl_text=owl_text.replace("erfÃ¼lltDurch","erfülltDurch") \
    .replace("erfÃ¼lltZiel","erfülltZiel")
    if not owl_text.startswith("<?xml"):
        print("Überprüfen Sie die Ontologie")
    return owl_text

#Umwandlung des ontology_text in eine Ontologie
def load_ontology_text(owl_text:str):
    #Schaffung der Umgebung
    owl_container=World()

    #Erstellung eines temporären Speicherorts für die Ontologie
    with tempfile.NamedTemporaryFile(delete=False, suffix=".owl",mode= "w" ,encoding= "utf-8") as vtemp:
        vtemp.write(owl_text)
        temp_path=vtemp.name
        onto=owl_container.get_ontology(temp_path).load()

        return onto 


def instanziierung(onto,df,projekt_name:str,output_path:str):

    
    #Projektname ohne Leerzeichen
    neu_projekt=projekt_name.replace(" ","_")

    #Klasse
    Forschungsprojekt=onto.Forschungsprojekt
 
    #Es wird überprüft, ob das Projekt bereits in der Ontologie vorhanden ist
    projekt=onto[neu_projekt]

    if projekt is None:
        projekt=Forschungsprojekt(neu_projekt)
        projekt.label=[projekt_name]

    pop_erfuellt=onto.world.search_one(iri="*#erfülltDurch")
    
    if pop_erfuellt is None:
        raise ValueError("erfüllt Durch existiert nicht")
    

    ziel_zu_praktik={}
    rg_zu_ziele={}
    reifegrad_pro_phase={}

    for _, row in df.iterrows():
        phase= str(row["Phase"]).strip()
        ziel_index= row["Zielindex"]
        reifegrad=str(row["Reifegrad"]).strip()
        praktik_index=row["Praktikindex"]


        if reifegrad:
            num=int("".join(filter(str.isdigit, reifegrad)))
            reifegrad_pro_phase.setdefault(phase,set()).add(num)



        if not ziel_index or not reifegrad:
            continue

        #Index des Ziels erstellen
        ziel_id= "Z"+str(ziel_index).replace(".","_")
        
        #Ziel in der Ontologie in Unterklassen unterteilen (technisch/methodisch) und erstellen oder wiederverweden
        ziel=onto[ziel_id] 
        
        
        if ziel is None:
            
            zielbereich_typ=str(row["Zielbereich"]).strip().lower()

            if "technisch" in zielbereich_typ:
                ziel=onto.Technisches_Ziel(ziel_id)

            else:
                ziel=onto.Methodisches_Ziel(ziel_id)
        #ziel.comment.append(str(row["Ziel"]))
        ziel_text= str(row["Ziel"]).strip()
        if ziel_text and ziel_text not in ziel.comment:
            ziel.comment.append(ziel_text)

    

        #Verbindung von Prozessbereich mit Ziel
        pb=onto[f"{phase}_Indiv"]

        if pb is None:
            phase2=phase.replace(" ","").replace("-","_")
            pb=onto[f"{phase2}_Indiv"]
        if pb is None:
            continue    
        
        if ziel not in pb.charakterisiertDurch:
            pb.charakterisiertDurch.append(ziel)
        
        #Verbindung von Reifegrad mit Ziel
        rg=onto[f"{phase}_{reifegrad}_Indiv"]

        
        
        if rg is None:
            phase2=phase.replace(" ","").replace("-","_")
            rg=onto[f"{phase2}_{reifegrad}_Indiv"]
        if rg is None:
            continue
        if ziel not in pop_erfuellt[rg]:
            pop_erfuellt[rg].append(ziel)
        rg_key=(phase,reifegrad)
        rg_zu_ziele.setdefault(rg_key,set()).add(ziel)
        
        
        #Index der Praktik erstellen
        if praktik_index:
    
            praktik_id="P"+str(praktik_index).replace(".","_")


            praktik=onto[praktik_id] 
            
            if praktik is None:
                praktik=onto.Praktik(praktik_id)
            praktik.comment.append(str(row["Praktik"]))

            #Verbindung von Ziel mit Praktik
            if praktik not in ziel.erreichtDurch:
                ziel.erreichtDurch.append(praktik)
            ziel_zu_praktik.setdefault(ziel,set()).add(praktik)
        
    
    #SWRL erstellen
    #SWRL1 (Projekt,Ziel,Praktik)
    for ziel, praktik in ziel_zu_praktik.items():
        
        body=" ^ ".join([f"wendenPraktikAn(?proj,{p.name})" for p in praktik])
        head=f"erfülltZiel(?proj,{ziel.name})"
        ruletext=(f"{body}->{head}")

        with onto:
            rule=Imp()
            rule.set_as_rule(ruletext)
    
    #SWRL2(Projekt,Reifegrad,Phase)

    for (phase,reifegrad), ziele in rg_zu_ziele.items():

        reifegrad_nmr=int("".join(filter(str.isdigit,reifegrad)))
        ph_rg_indiv=onto[f"{phase}_{reifegrad}_Indiv"]
        if ph_rg_indiv is None:
            phase2=phase.replace(" ","").replace("-","_")
            ph_rg_indiv=onto[f"{phase2}_{reifegrad}_Indiv"]
        
        if ph_rg_indiv is None:
            continue
        #body= " ^ ".join([f"erfülltZiel(?proj,{z.name})" for z in ziele])

        body_teile=[f"erfülltZiel(?proj,{z.name})" for z in ziele]

        if (reifegrad_nmr-1) in reifegrad_pro_phase.get(phase, set()):

            vor_reifegrad_nmr= f"{phase}_Reifegrad{reifegrad_nmr-1}_Indiv"

            vor_reifegrad_nmr=onto[vor_reifegrad_nmr]

            if vor_reifegrad_nmr is None:
                phase2= phase.replace(" ", "").replace("-","_")

                vor_reifegrad_nmr=onto[f"{phase2}_Reifegrad{reifegrad_nmr-1}_Indiv"]
            
            if vor_reifegrad_nmr:
                body_teile.append(f"erzieltReifegrad(?proj,{vor_reifegrad_nmr.name})")
        body=" ^ ".join(body_teile)

        head= f"erzieltReifegrad(?proj,{ph_rg_indiv.name})"
        ruletext=f"{body}->{head}"

        with onto:
            rule=Imp()
            rule.set_as_rule(ruletext)
    
  
        
        

    onto.save(file=output_path,format="rdfxml")
        


 
    return projekt

#liest die Praktikindex und Zielindex aus dem Excel-Blatt 'Bewertungdata' mit angekreuzte Checkbox
def read_bewertungdata(xlsm_path,phase):

    #Offnung von excel, Zugriff auf das Blatt BewertungData und Lesen
    wb=load_workbook(xlsm_path,data_only=True,keep_vba=True)
    sheet=wb["BewertungData"]
    rows=list(sheet.iter_rows(values_only=True))

    #Wenn nur überschrift vorhanden ist (leeres Blatt)=keine Auswahl
    if len(rows)<2:
        return[]
    auswahl=[]
    #Durchläuft die Reihen der Excel
    for zeile in rows[1:]:
        phase_zelle, zielindex, praktikindex=zeile

        if praktikindex is None:
            continue 
        auswahl.append((phase_zelle,zielindex,praktikindex))
            
    return auswahl


#Anwendung der Auswahl auf die Ontologie
def bewertung_checkboxen(onto, projekt_indiv,auswahl):
    #Sucht die Eigenschaft "wendenPraktikAn"
    propert=onto.world.search_one(iri="*#wendenPraktikAn")

    propert[projekt_indiv].clear()

    

    #Durchlauf der Praktiken und Ziele
    for _, _, praktikindex in auswahl:
        praktikindex="P" + str(praktikindex).replace(".","_")
        praktik=onto[praktikindex]

        if praktik is  None:
            continue

        propert[projekt_indiv].append(praktik)

        

    

def reasoner_reifegrad(onto, projekt_indiv,phase):

    

    #Sucht die Eigenschaft "erzieltReifegrad"
    propert=onto.world.search_one(iri="*#erzieltReifegrad")

    for rg in list(propert[projekt_indiv]):
        if rg.name.startswith(phase + "_"):
            propert[projekt_indiv].remove(rg)

    #Reasoner Pellet ausführen

    onto.world._reasoning=False
    sync_reasoner_pellet(infer_property_values=True, infer_data_property_values=True)
    
    
    #Erhalt aller Reifegrade
    reifegrade=list(propert[projekt_indiv])

    #nur diejenigen der entsprechenden Phase gefiltert
    phase_name=phase+ "_"
    phase_reifegrad=[rg for rg in reifegrade if rg.name.startswith(phase_name)]

    if not phase_reifegrad:
        return "Reifegrad nicht erreicht"
    
    return sorted(phase_reifegrad,key=lambda x:x.name)[-1].name

def write_erzielteReifegrad_excel(xlsm_path, phase, reifegrad,auswahl):

    if reifegrad != "Reifegrad nicht erreicht":
        reifegrad= reifegrad.split("Reifegrad")[-1].split("_")[0]

    wb=xw.Book(xlsm_path)
    sheet=wb.sheets[phase]

    #Uberschrift
    sheet.range("A20").value="Reifegrad: "
    sheet.range("A20").api.Font.Size=13
    sheet.range("A20").api.Font.Bold=True

    #Wert
    sheet.range("B20").value=reifegrad
    sheet.range("B20").api.Font.Size=12
    sheet.range("B20").api.Font.Bold=False


    #die nächste erfüllende Praktik
    selected=[t[2] for t in auswahl if t[0].lower()==phase.lower()]
    row=31

    neu_text= ""
    while True:
        praktik_index=sheet.range(f"C{row}").value
        praktik_text=sheet.range(f"D{row}").value

        if not praktik_index:
            sheet.range("B21").value="erfüllt"
            neu_text= "erfüllt"
            break

        if str(praktik_index) not in selected:
            sheet.range("A21").value= "Nächste:"
            sheet.range("A21").api.Font.Bold=True
            sheet.range("A21").api.Font.Size=13
            sheet.range("B21").value=praktik_text
            sheet.range("B21").api.Font.Size=12
            neu_text=praktik_text
            break
    
        row +=1
    
    #Die Informationen werden im Blatt „Gesamtbild“ gespeichert.
    if "Gesamtbild" in [s.name for s in wb.sheets]:
        gb=wb.sheets["Gesamtbild"]
        zeile=4

        while True:
            ph= gb.range(f"A{zeile}").value
            if ph is None or str(ph).strip() =="":
                break

            if str(ph).strip().lower()== str(phase).strip().lower():
                gb.range(f"B{zeile}").value= reifegrad
                gb.range(f"G{zeile}").value= neu_text
                

                break
            zeile +=1



   



    



    


def main():

    if len(sys.argv)>=5 and sys.argv[4].lower()=="reason":  #if len(sys.argv)>=5 and sys.argv[4].lower()=="reason":
        xlsm_container= Path(sys.argv[1])
        projekt_name=sys.argv[2]
        phase=sys.argv[3]
        

        print("Phase ist: ",phase)

        owl_path=xlsm_container.parent/f"{projekt_name}.owl"  
        

        print("owl exists:", owl_path.exists())

        onto_path.append(str(owl_path.parent))
        onto=get_ontology(owl_path.name).load()
        projekt_indiv = onto[projekt_name.replace(" ", "_")]
        print("Mein Projekt ist: ", projekt_indiv)

        auswahl=read_bewertungdata(xlsm_container,phase)
        print("Der Auswahl ist:", auswahl)

        bewertung_checkboxen(onto,projekt_indiv,auswahl)
        onto.save(file=str(owl_path), format="rdfxml")
       

       #para solucionar el problema

        reifegrad=reasoner_reifegrad(onto,projekt_indiv,phase)
        print("reifegrad", reifegrad)
        write_erzielteReifegrad_excel(xlsm_container,phase,reifegrad,auswahl)

        return

    #überprüfung der Erfassung der drei erwarteten Argumente aus Excel ("ImportedData", "Ontology" und ForschungsprojektName)
    if len(sys.argv) < 4:
        raise SystemExit("python Instanziierungmain.py <excel_import> <xslm_container> <projekt_name>")
    
    #importierte Pfade aus Excel
    excel_import=Path(sys.argv[1])
    xlsm_container=Path(sys.argv[2])
    projekt_name=sys.argv[3]
    

    output_path=str(xlsm_container.parent/f"{projekt_name}.owl")  
    

    #korrekte Importprüfung
    print("Weg Excel import:",excel_import)
    print("xlsm container:", xlsm_container)
    print("Name des Froschungsprojekts:",projekt_name)


    df= read_imported_data(xlsm_container)

    #print("Phase counts:")
    #print(df["Phase"].value_counts())
    write_sheets(str(xlsm_container), df, first_row=30)


    #print([repr(x) for x in df["Phase"].unique()])
    print("ImportedData lines:", len(df))
    print("ImportedData columns:", list(df.columns))


    

    owl_text= read_ontology_text(xlsm_container)
    print("Ontology text:", len(owl_text))

    onto=load_ontology_text(owl_text)
    print("Ontology:")
    print(onto.base_iri)
    

    
    

   

    instanziierung(onto,df,projekt_name,output_path)
    
    
    


if __name__=="__main__":
    main()

